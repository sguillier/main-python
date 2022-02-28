import pandas as pd

array_test = ['AED', 'ARS', 'AUD', 'BMD', 'BRL', 'CAD', 'CNF', 'NO', 'CNY', 'COP',
                                'CZK', 'DKK', 'DOP', 'EGP', 'EUR', 'GBP', 'HKD', 'HUF', 'IDR', 'ILS',
                                'INR', 'JPY', 'KRW', 'KYD', 'KZT', 'MXN', 'MYR', 'NOK', 'NZD', 'PAB',
                                'PEN', 'PHP', 'PLN', 'RUB', 'SEK', 'SGD', 'THB', 'TRY', 'TWD', 'UAH',
                                'US$', 'VEF', 'ZAR', 'UF', 'OT']

for element in array_test:
    print(element)



from openpyxl import load_workbook
from openpyxl.styles import Alignment

#Nombre del libro
report_name='excel_test.xlsx'
wb = load_workbook(filename=report_name)

#Se puede llamar a la hoja del libro de esta forma
ws = wb['sheet_test']


#Agrega a la celda 1,1 el valor 'valor_test'
row_index = 1

ws.cell(row=row_index, column=1).value = 'valor_test'

#Le da formato a la celda
ws.cell(row=row_index, column=1).alignment = Alignment(horizontal="left", vertical="top")

#Guarda y cierra la conexion
wb.save(report_name)
wb.close()


#diccionarios y excel, inserta en la hoja sheet_test_2 los valores del diccionario
row_index = 2
ws = wb['sheet_test_2']

for element in array_test:
    ws.cell(row=row_index, column=1).value =element
    row_index+=1
wb.save(report_name)
wb.close()


#Lee el excel y lo guarda en un df y lo recorre
print('_________________Lee el excel y lo guarda en un df y lo recorre_______________________')

df=pd.read_excel('excel_test.xlsx',sheet_name='sheet_test_3')

#Recorre el df 
for index, row in df.iterrows():
    print(row['column_test'])

print('_________________transformar los elementos del df en una arreglo_______________________')

#transforma los elements del df en una arreglo

array_test_2=df['column_test'].unique()

for lista in array_test_2:
    print(lista)



#recorre las filas del DF
print('___________recorre las filas del DF_____________________________')
for lista in df['column_test'].keys():
    print(lista)

