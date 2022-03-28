import openpyxl

book = openpyxl.load_workbook('test.xlsx')

hoja = book.active

from openpyxl.styles import PatternFill, Font
hoja['A1'] = 'Hello World 2'

hoja['A4'].fill = PatternFill('solid', fgColor='ff0000') # rojo
hoja['A5'].fill = PatternFill('solid', fgColor='ffc000') # naranja
hoja['A6'].fill = PatternFill('solid', fgColor='ffff00') # amarillo
hoja['A7'].fill = PatternFill('solid', fgColor='00ff00') # verde
hoja['A8'].fill = PatternFill('solid', fgColor='0000ff') # azul
hoja['A9'].fill = PatternFill('solid', fgColor='ff00ff') # violeta
hoja['A10'].fill = PatternFill('solid', fgColor='00ffff') # celeste


hoja['A4'].font = Font(color='FFFFFFFF') # blanco
hoja['A5'].font = Font(color='000FFF00') # verde
hoja['A6'].font = Font(color='000FFFF0') # celeste




celdas = hoja['A1':'C3']

for fila in celdas:
    for celda in fila:
        print(celda.value)


book.save('test.xlsx')
