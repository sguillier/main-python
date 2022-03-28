

import openpyxl
from getRutByNombre import getRutByNombre
from getNombreByRut import getNombreByRut


nombre = getNombreByRut('18912988-2')
print(nombre)
# nombre = getNombreByRut('76163495-K')

# path = 'listado.xlsx'

# book = openpyxl.load_workbook(path)
# hoja = book.active

# for i in range(2, hoja.max_row+1):
#     # celda = hoja.cell(row=i, column=1)
#     rut = hoja.cell(row=i, column=2).value
#     print(i, ' ', rut)
#     nombre = getNombreByRut(rut)
#     print(nombre)
#     hoja.cell(row=i, column=3).value = nombre


# book.save(path)
