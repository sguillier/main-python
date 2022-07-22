

from asyncio.windows_events import NULL
from random import randint
from numpy import empty
import openpyxl
from getRutByNombre import getRutByNombre
from getNombreByRut import getNombreByRut
import time

# rut = getRutByNombre('CHIPTIME PRODUCCIONES SPA')
# print(rut)

path = 'listado.xlsx'

book = openpyxl.load_workbook(path)
hoja = book.active

for i in range(2, 1000):
    nombre = hoja.cell(row=i, column=1).value
    if str(nombre) == 'None':
        break
    t = randint(1,8)
    time.sleep(t)
    rut = getRutByNombre(nombre)
    print(i, ' ', nombre,' ',t, ' ', rut)
    hoja.cell(row=i, column=2).value = rut


book.save(path)
